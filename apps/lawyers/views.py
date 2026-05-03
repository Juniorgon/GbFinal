"""
apps/lawyers/views.py
---------------------
Cadastro de advogados: apenas administradores.
Verificacao real no backend via @admin_required.
"""

import io

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import connection, DatabaseError, transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render

from apps.accounts.models import CustomUser
from apps.accounts.permissions import admin_required, get_branch_object_or_403
from apps.branches.models import Branch

from .models import Lawyer


def get_branch(request):
    return getattr(request, 'current_branch', None) or request.user.branch


def split_full_name(full_name):
    parts = full_name.split(None, 1)
    first_name = parts[0] if parts else ''
    last_name = parts[1] if len(parts) > 1 else ''
    return first_name, last_name


def build_form_context(title, lawyer=None, form_data=None):
    linked_user = getattr(lawyer, 'user', None)
    all_branches = Branch.objects.filter(is_active=True).order_by('-is_headquarters', 'name')
    selected_branch_count = 0
    if form_data:
        selected_branch_count = all_branches.count() if form_data.get('allow_all_branches') else 1
    elif lawyer:
        selected_ids = set(lawyer.accessible_branches.values_list('id', flat=True))
        if lawyer.branch_id:
            selected_ids.add(lawyer.branch_id)
        selected_branch_count = len(selected_ids)

    return {
        'title': title,
        'lawyer': lawyer,
        'estado_choices': Lawyer.ESTADO_CHOICES,
        'linked_user': linked_user,
        'form_data': form_data,
        'all_branches': all_branches,
        'allow_all_branches': selected_branch_count >= all_branches.count() and all_branches.count() > 1,
    }


def validate_passwords(password, confirm_password, require_password=False):
    if require_password and not password:
        raise ValueError('Informe a senha de acesso do advogado.')
    if password != confirm_password:
        raise ValueError('Senha e confirmacao de senha nao conferem.')


def ensure_legacy_lawyer_defaults():
    """
    Compatibilidade para bancos ainda nao migrados apos a remocao de campos
    antigos. Se a coluna existir, garante default para inserts novos.
    """
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                'ALTER TABLE lawyers_lawyer '
                'ALTER COLUMN can_access_financial SET DEFAULT false'
            )
    except DatabaseError:
        pass


def sync_lawyer_branch_access(lawyer, allow_all_branches):
    if allow_all_branches:
        lawyer.accessible_branches.set(Branch.objects.filter(is_active=True))
    else:
        lawyer.accessible_branches.clear()


@login_required
def lawyer_list(request):
    branch = get_branch(request)
    qs = Lawyer.objects.filter(branch=branch).order_by('name')
    total = qs.count()
    ativos = qs.filter(is_active=True).count()
    estados = qs.filter(is_active=True).values_list('oab_state', flat=True).distinct().count()
    return render(request, 'lawyers/list.html', {
        'lawyers': qs, 'total': total, 'ativos': ativos, 'estados': estados,
        'estado_choices': Lawyer.ESTADO_CHOICES,
        'all_branches': Branch.objects.filter(is_active=True).order_by('-is_headquarters', 'name'),
    })


@login_required
@admin_required
def lawyer_create(request):
    """Apenas admins podem cadastrar advogados."""
    branch = get_branch(request)

    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            email = request.POST.get('email', '').strip()
            username = request.POST.get('username', '').strip()
            password = request.POST.get('password', '')
            confirm_password = request.POST.get('confirm_password', '')
            phone = request.POST.get('phone', '').strip()

            if not username:
                raise ValueError('Informe o nome de usuario para login.')
            validate_passwords(password, confirm_password, require_password=True)
            if CustomUser.objects.filter(username=username).exists():
                raise ValueError('Ja existe um usuario com esse login.')

            first_name, last_name = split_full_name(name)
            ensure_legacy_lawyer_defaults()
            with transaction.atomic():
                user = CustomUser.objects.create_user(
                    username=username,
                    password=password,
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    role=CustomUser.ROLE_LAWYER,
                    branch=branch,
                    phone=phone,
                )
                lawyer = Lawyer.objects.create(
                    branch=branch,
                    user=user,
                    name=name,
                    email=email,
                    oab_number=request.POST.get('oab_number', '').strip(),
                    oab_state=request.POST.get('oab_state', 'RS'),
                    phone=phone,
                    specialization=request.POST.get('specialization', '').strip(),
                    is_active=True,
                )
                sync_lawyer_branch_access(lawyer, 'allow_all_branches' in request.POST)
            messages.success(request, f'Advogado {lawyer.name} registrado com sucesso!')
            return redirect('lawyers:list')
        except Exception as e:
            messages.error(request, f'Erro ao registrar advogado: {e}')
            return render(request, 'lawyers/form.html', build_form_context(
                'Registrar Advogado',
                form_data=request.POST,
            ))

    return render(request, 'lawyers/form.html', build_form_context(
        'Registrar Advogado',
    ))


@login_required
@admin_required
def lawyer_edit(request, pk):
    """Apenas admins podem editar advogados."""
    lawyer = get_branch_object_or_403(Lawyer, request, pk=pk)

    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            email = request.POST.get('email', '').strip()
            username = request.POST.get('username', '').strip()
            password = request.POST.get('password', '')
            confirm_password = request.POST.get('confirm_password', '')
            phone = request.POST.get('phone', '').strip()
            first_name, last_name = split_full_name(name)

            with transaction.atomic():
                if not username:
                    raise ValueError('Informe o nome de usuario para login.')
                if password or confirm_password:
                    validate_passwords(password, confirm_password)
                username_qs = CustomUser.objects.filter(username=username)
                if lawyer.user:
                    username_qs = username_qs.exclude(pk=lawyer.user_id)
                if username_qs.exists():
                    raise ValueError('Ja existe um usuario com esse login.')

                if lawyer.user:
                    lawyer.user.username = username
                    lawyer.user.email = email
                    lawyer.user.first_name = first_name
                    lawyer.user.last_name = last_name
                    lawyer.user.phone = phone
                    lawyer.user.branch = lawyer.branch
                    lawyer.user.role = CustomUser.ROLE_LAWYER
                    if password:
                        lawyer.user.set_password(password)
                    lawyer.user.save()
                else:
                    validate_passwords(password, confirm_password, require_password=True)
                    lawyer.user = CustomUser.objects.create_user(
                        username=username,
                        password=password,
                        email=email,
                        first_name=first_name,
                        last_name=last_name,
                        role=CustomUser.ROLE_LAWYER,
                        branch=lawyer.branch,
                        phone=phone,
                    )

                lawyer.name = name
                lawyer.email = email
                lawyer.oab_number = request.POST.get('oab_number', '').strip()
                lawyer.oab_state = request.POST.get('oab_state', 'RS')
                lawyer.phone = phone
                lawyer.specialization = request.POST.get('specialization', '').strip()
                lawyer.save()
                sync_lawyer_branch_access(lawyer, 'allow_all_branches' in request.POST)
            messages.success(request, 'Advogado atualizado!')
            return redirect('lawyers:list')
        except Exception as e:
            messages.error(request, f'Erro ao atualizar advogado: {e}')
            return render(request, 'lawyers/form.html', build_form_context(
                'Editar Advogado',
                lawyer=lawyer,
                form_data=request.POST,
            ))

    return render(request, 'lawyers/form.html', build_form_context(
        'Editar Advogado',
        lawyer=lawyer,
    ))


@login_required
@admin_required
def lawyer_toggle(request, pk):
    """Apenas admins podem ativar/desativar."""
    lawyer = get_branch_object_or_403(Lawyer, request, pk=pk)
    lawyer.is_active = not lawyer.is_active
    lawyer.save(update_fields=['is_active'])
    if lawyer.user:
        lawyer.user.is_active = lawyer.is_active
        lawyer.user.save(update_fields=['is_active'])
    status = 'ativado' if lawyer.is_active else 'desativado'
    messages.success(request, f'Advogado {status}.')
    return redirect('lawyers:list')


@login_required
@admin_required
def lawyer_delete(request, pk):
    """Apenas admins podem excluir advogados."""
    if request.method != 'POST':
        return redirect('lawyers:list')

    lawyer = get_branch_object_or_403(Lawyer, request, pk=pk)
    if lawyer.user_id == request.user.id:
        messages.error(request, 'Voce nao pode excluir seu proprio usuario por aqui.')
        return redirect('lawyers:list')

    lawyer_name = lawyer.name
    linked_user = lawyer.user
    with transaction.atomic():
        lawyer.accessible_branches.clear()
        lawyer.delete()
        if linked_user:
            linked_user.delete()

    messages.success(request, f'Advogado {lawyer_name} excluido com sucesso.')
    return redirect('lawyers:list')


@login_required
def export_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    branch = get_branch(request)
    lawyers = Lawyer.objects.filter(branch=branch)
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    w, h = A4
    p.setFillColor(colors.HexColor('#1a1a2e'))
    p.rect(0, h - 60, w, 60, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont('Helvetica-Bold', 14)
    p.drawString(40, h - 38, 'GB & N.Comin Advocacia - Advogados')
    y = h - 90
    p.setFillColor(colors.HexColor('#e8650a'))
    p.setFont('Helvetica-Bold', 9)
    for txt, x in [('Nome', 40), ('OAB', 220), ('Email', 290), ('Telefone', 420), ('Status', 500)]:
        p.drawString(x, y, txt)
    y -= 20
    p.setFont('Helvetica', 8)
    for lawyer in lawyers:
        if y < 60:
            p.showPage()
            y = h - 60
        p.setFillColor(colors.black)
        p.drawString(40, y, lawyer.name[:28])
        p.drawString(220, y, lawyer.oab_display)
        p.drawString(290, y, lawyer.email[:22])
        p.drawString(420, y, lawyer.phone)
        p.drawString(500, y, 'Ativo' if lawyer.is_active else 'Inativo')
        y -= 16
    p.save()
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="advogados.pdf"'
    return response


@login_required
def export_excel(request):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    branch = get_branch(request)
    lawyers = Lawyer.objects.filter(branch=branch)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Advogados'
    headers = ['Nome', 'OAB', 'Email', 'Telefone', 'Especializacao', 'Status']
    hf = PatternFill(start_color='E8650A', end_color='E8650A', fill_type='solid')
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = hf
        cell.alignment = Alignment(horizontal='center')
    for lawyer in lawyers:
        ws.append([
            lawyer.name,
            lawyer.oab_display,
            lawyer.email,
            lawyer.phone,
            lawyer.specialization or 'Nao informado',
            'Ativo' if lawyer.is_active else 'Inativo',
        ])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="advogados.xlsx"'
    return response
