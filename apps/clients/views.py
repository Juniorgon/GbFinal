from datetime import date
import io
from textwrap import wrap

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from apps.accounts.permissions import _user_can_access_branch
from apps.branches.models import Branch

from .models import Client


def get_branch(request):
    current_branch = getattr(request, "current_branch", None)
    if current_branch and _user_can_access_branch(request.user, current_branch):
        return current_branch

    lawyer_profile = getattr(request.user, "lawyer_profile", None)
    if lawyer_profile and lawyer_profile.branch and lawyer_profile.branch.is_active:
        return lawyer_profile.branch

    user_branch = getattr(request.user, "branch", None)
    if user_branch and _user_can_access_branch(request.user, user_branch):
        return user_branch

    if request.user.is_super_admin:
        return Branch.get_default_branch()

    return None


def build_client_address(client):
    parts = []
    if client.street:
        street_line = client.street
        if client.number:
            street_line += f", {client.number}"
        if client.complement:
            street_line += f" - {client.complement}"
        parts.append(street_line)
    if client.district:
        parts.append(client.district)
    if client.city or client.state:
        parts.append(f"{client.city or 'Cidade nao informada'}/{client.state or 'RS'}")
    if client.zipcode:
        parts.append(f"CEP {client.zipcode}")
    return ", ".join(parts) or "endereco nao informado"


def build_procuracao_paragraphs(client, branch):
    address = build_client_address(client)
    branch_address = branch.address if branch and branch.address else 'endereco profissional nao informado'
    branch_city = f"{branch.city or ''}/{branch.state or ''}".strip('/').strip() if branch else ''
    office_location = branch_address + (f", {branch_city}." if branch_city else ".")
    city_line = branch.city or client.city or 'Cidade'
    current_date = date.today().strftime('%d/%m/%Y')

    if client.type == Client.TYPE_JURIDICA:
        outorgante = (
            f"{client.name}, pessoa juridica de direito privado, inscrita no CNPJ sob o n. "
            f"{client.cpf_cnpj or 'nao informado'}, com sede em {address}."
        )
        return [
            "OUTORGANTE:",
            outorgante,
            "",
            "OUTORGADO(S): GB & N.Comin Advocacia, por seus advogados regularmente inscritos na OAB, com escritorio profissional em "
            + office_location,
            "",
            "PODERES: o presente instrumento confere poderes da clausula ad judicia et extra para representar a outorgante em processos judiciais, administrativos e extrajudiciais, "
            "podendo propor acoes, apresentar defesa, acompanhar procedimentos, firmar acordos, receber valores, dar quitacao, substabelecer com ou sem reserva de poderes e praticar todos os atos necessarios ao fiel cumprimento deste mandato.",
            "",
            "Ficam igualmente conferidos poderes especiais para confessar, reconhecer a procedencia do pedido, transigir, desistir, renunciar ao direito sobre o qual se funda a acao, receber, dar quitacao e firmar compromissos, quando isso atender aos interesses da outorgante.",
            "",
            f"{city_line}, {current_date}.",
            "",
            "",
            "___________________________________",
            client.name,
            "Outorgante",
        ]

    nationality = client.nationality or 'brasileiro(a)'
    marital_status = client.marital_status or 'estado civil nao informado'
    profession = client.profession or 'profissao nao informada'
    outorgante = (
        f"{client.name}, {nationality}, {marital_status}, {profession}, portador(a) do CPF n. "
        f"{client.cpf_cnpj or 'nao informado'}, residente e domiciliado(a) em {address}."
    )
    return [
        "OUTORGANTE:",
        outorgante,
        "",
        "OUTORGADO(S): GB & N.Comin Advocacia, por seus advogados regularmente inscritos na OAB, com escritorio profissional em "
        + office_location,
        "",
        "PODERES: o presente instrumento confere poderes da clausula ad judicia et extra para o foro em geral, em qualquer juizo, instancia ou tribunal, "
        "podendo propor acoes, acompanhar processos, apresentar defesa, recorrer, transigir, desistir, receber valores, dar quitacao, substabelecer com ou sem reserva de poderes e praticar todos os atos necessarios ao fiel cumprimento deste mandato.",
        "",
        "Ficam ainda outorgados poderes especiais para confessar, reconhecer a procedencia do pedido, firmar compromissos, celebrar acordos e receber citacoes e intimacoes, sempre em defesa dos interesses do(a) outorgante.",
        "",
        f"{city_line}, {current_date}.",
        "",
        "",
        "___________________________________",
        client.name,
        "Outorgante",
    ]


def escape_pdf_text(text):
    return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def pdf_text(x, y, text, size=11, font='F1', color=(0, 0, 0)):
    r, g, b = color
    return [
        "BT",
        f"/{font} {size} Tf",
        f"{r:.3f} {g:.3f} {b:.3f} rg",
        f"1 0 0 1 {x} {y} Tm",
        f"({escape_pdf_text(text)}) Tj",
        "ET",
    ]


def split_procuracao_paragraphs(paragraphs, client_name="", width=82):
    blocks = []
    for paragraph in paragraphs:
        if not paragraph:
            blocks.append({"type": "spacer", "height": 12})
            continue

        if paragraph == "OUTORGANTE:":
            blocks.append({
                "type": "lines",
                "lines": [paragraph],
                "font": "F2",
                "size": 11,
                "color": (0.10, 0.18, 0.26),
                "line_height": 16,
                "x": 56,
            })
            continue

        if paragraph.startswith("OUTORGADO(S):") or paragraph.startswith("PODERES:"):
            wrapped_lines = wrap(paragraph, width=width)
            blocks.append({
                "type": "lines",
                "lines": wrapped_lines,
                "font": "F1",
                "size": 11,
                "color": (0.35, 0.39, 0.45),
                "line_height": 17,
                "x": 56,
            })
            blocks.append({"type": "spacer", "height": 6})
            continue

        if paragraph == "___________________________________":
            blocks.append({"type": "spacer", "height": 18})
            blocks.append({
                "type": "lines",
                "lines": [paragraph],
                "font": "F1",
                "size": 11,
                "color": (0.35, 0.39, 0.45),
                "line_height": 14,
                "x": 190,
            })
            continue

        if paragraph == client_name or paragraph == "Outorgante":
            blocks.append({
                "type": "lines",
                "lines": [paragraph],
                "font": "F1" if paragraph == client_name else "F2",
                "size": 11 if paragraph == client_name else 10,
                "color": (0.35, 0.39, 0.45),
                "line_height": 15,
                "x": 220 if paragraph == client_name else 248,
            })
            continue

        wrapped_lines = wrap(paragraph, width=width)
        blocks.append({
            "type": "lines",
            "lines": wrapped_lines,
            "font": "F1",
            "size": 11,
            "color": (0.35, 0.39, 0.45),
            "line_height": 17,
            "x": 70,
        })
        blocks.append({"type": "spacer", "height": 8})

    return blocks


def draw_pdf_header(content_lines, page_number, navy, gold, light):
    content_lines.extend([
        "q",
        f"{navy[0]:.3f} {navy[1]:.3f} {navy[2]:.3f} rg",
        "0 760 595 82 re f",
        "Q",
        "q",
        f"{navy[0]:.3f} {navy[1]:.3f} {navy[2]:.3f} rg",
        "0 0 595 54 re f",
        "Q",
        "q",
        f"{gold[0]:.3f} {gold[1]:.3f} {gold[2]:.3f} RG",
        "2.2 w",
        "56 790 m 70 804 l S",
        "70 804 m 84 790 l S",
        "70 804 m 70 776 l S",
        "48 786 m 92 786 l S",
        "56 786 m 52 774 l S",
        "84 786 m 88 774 l S",
        "38 786 m 52 786 l S",
        "88 786 m 102 786 l S",
        "Q",
    ])
    content_lines.extend(pdf_text(112, 796, "GB & n.comin", size=23, font='F2', color=gold))
    content_lines.extend(pdf_text(112, 772, "advocacia", size=16, font='F2', color=gold))
    content_lines.extend(pdf_text(405, 788, "PROCURACAO", size=12, font='F2', color=light))
    content_lines.extend(pdf_text(390, 772, "Ad judicia et extra", size=9, font='F1', color=light))
    if page_number == 1:
        content_lines.extend(pdf_text(56, 730, "Instrumento particular de mandato", size=15, font='F2', color=navy))


def draw_pdf_footer(content_lines, client_name, branch_name, page_number, gold, light):
    footer_left = branch_name or "GB & N.Comin Advocacia"
    footer_right = client_name or "Documento"
    content_lines.extend(pdf_text(56, 22, footer_left, size=9, font='F2', color=gold))
    content_lines.extend(pdf_text(355, 22, footer_right, size=9, font='F1', color=light))
    content_lines.extend(pdf_text(520, 22, str(page_number), size=9, font='F1', color=light))


def build_simple_pdf(paragraphs, client_name='', branch_name=''):
    navy = (0.10, 0.18, 0.26)
    gold = (0.86, 0.66, 0.42)
    light = (0.95, 0.91, 0.84)
    blocks = split_procuracao_paragraphs(paragraphs, client_name=client_name, width=82)
    top_margin = 702
    bottom_limit = 86

    page_contents = []
    page_number = 1
    content_lines = []
    draw_pdf_header(content_lines, page_number, navy, gold, light)
    current_y = top_margin

    for block in blocks:
        if block["type"] == "spacer":
            current_y -= block["height"]
            continue

        block_height = len(block["lines"]) * block["line_height"]
        if current_y - block_height < bottom_limit:
            draw_pdf_footer(content_lines, client_name, branch_name, page_number, gold, light)
            page_contents.append("\n".join(content_lines).encode("latin-1", errors="replace"))
            page_number += 1
            content_lines = []
            draw_pdf_header(content_lines, page_number, navy, gold, light)
            current_y = 720

        for line in block["lines"]:
            content_lines.extend(
                pdf_text(
                    block["x"],
                    current_y,
                    line,
                    size=block["size"],
                    font=block["font"],
                    color=block["color"],
                )
            )
            current_y -= block["line_height"]

    draw_pdf_footer(content_lines, client_name, branch_name, page_number, gold, light)
    page_contents.append("\n".join(content_lines).encode("latin-1", errors="replace"))

    objects = []
    objects.append(b"<< /Type /Catalog /Pages 2 0 R >>")

    page_object_numbers = []
    content_object_numbers = []
    next_object_number = 3
    for _ in page_contents:
        page_object_numbers.append(next_object_number)
        content_object_numbers.append(next_object_number + 1)
        next_object_number += 2

    kids = " ".join(f"{page_number_obj} 0 R" for page_number_obj in page_object_numbers)
    objects.append(f"<< /Type /Pages /Kids [{kids}] /Count {len(page_contents)} >>".encode("latin-1"))

    for page_object_number, content_object_number, content in zip(page_object_numbers, content_object_numbers, page_contents):
        objects.append(
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Contents {content_object_number} 0 R /Resources << /Font << /F1 {next_object_number} 0 R /F2 {next_object_number + 1} 0 R >> >> >>".encode("latin-1")
        )
        objects.append(f"<< /Length {len(content)} >>\nstream\n".encode("latin-1") + content + b"\nendstream")

    objects.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    objects.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>")

    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(len(pdf))
        pdf.extend(f"{index} 0 obj\n".encode("latin-1"))
        pdf.extend(obj)
        pdf.extend(b"\nendobj\n")

    xref_start = len(pdf)
    pdf.extend(f"xref\n0 {len(objects) + 1}\n".encode("latin-1"))
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("latin-1"))
    pdf.extend(f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_start}\n%%EOF".encode("latin-1"))
    return bytes(pdf)


@login_required
def client_list(request):
    branch = get_branch(request)
    qs = Client.objects.filter(branch=branch, is_active=True)
    q = request.GET.get('q', '')
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(cpf_cnpj__icontains=q) | Q(phone__icontains=q))
    tipo = request.GET.get('tipo', '')
    if tipo:
        qs = qs.filter(type=tipo)
    return render(request, 'clients/list.html', {'clients': qs, 'q': q, 'tipo': tipo, 'branch': branch})


@login_required
def client_create(request):
    branch = get_branch(request)

    if not branch:
        messages.error(
            request,
            'Seu usuario nao esta vinculado a nenhuma filial. Vincule uma filial antes de cadastrar clientes.'
        )
        return redirect('clients:list')

    if request.method == 'POST':
        try:
            client = Client(
                branch=branch,
                name=request.POST.get('name', ''),
                type=request.POST.get('type', 'fisica'),
                cpf_cnpj=request.POST.get('cpf_cnpj', ''),
                nationality=request.POST.get('nationality', 'Brasileira'),
                marital_status=request.POST.get('marital_status', ''),
                profession=request.POST.get('profession', ''),
                phone=request.POST.get('phone', ''),
                email=request.POST.get('email', ''),
                street=request.POST.get('street', ''),
                number=request.POST.get('number', ''),
                complement=request.POST.get('complement', ''),
                district=request.POST.get('district', ''),
                city=request.POST.get('city', ''),
                state=request.POST.get('state', ''),
                zipcode=request.POST.get('zipcode', ''),
                notes=request.POST.get('notes', ''),
                created_by=request.user,
            )

            birth_date = request.POST.get('birth_date', '')
            if birth_date:
                client.birth_date = birth_date

            client.save()
            messages.success(request, f'Cliente {client.name} cadastrado com sucesso!')
            return redirect('clients:list')
        except Exception as e:
            messages.error(request, f'Erro ao cadastrar cliente: {e}')

    return render(request, 'clients/form.html', {'title': 'Novo Cliente', 'branch': branch})


@login_required
def client_edit(request, pk):
    branch = get_branch(request)
    client = get_object_or_404(Client, pk=pk, branch=branch)
    if request.method == 'POST':
        try:
            client.name = request.POST.get('name', client.name)
            client.type = request.POST.get('type', client.type)
            client.cpf_cnpj = request.POST.get('cpf_cnpj', '')
            client.nationality = request.POST.get('nationality', '')
            client.marital_status = request.POST.get('marital_status', '')
            client.profession = request.POST.get('profession', '')
            client.phone = request.POST.get('phone', '')
            client.email = request.POST.get('email', '')
            client.street = request.POST.get('street', '')
            client.number = request.POST.get('number', '')
            client.complement = request.POST.get('complement', '')
            client.district = request.POST.get('district', '')
            client.city = request.POST.get('city', '')
            client.state = request.POST.get('state', '')
            client.zipcode = request.POST.get('zipcode', '')
            client.notes = request.POST.get('notes', '')
            birth_date = request.POST.get('birth_date', '')
            if birth_date:
                client.birth_date = birth_date
            client.save()
            messages.success(request, 'Cliente atualizado com sucesso!')
            return redirect('clients:list')
        except Exception as e:
            messages.error(request, f'Erro: {e}')
    return render(request, 'clients/form.html', {'title': 'Editar Cliente', 'client': client, 'branch': branch})


@login_required
def client_delete(request, pk):
    branch = get_branch(request)
    client = get_object_or_404(Client, pk=pk, branch=branch)
    client.is_active = False
    client.save_without_historical_record()
    messages.success(request, f'Cliente {client.name} removido.')
    return redirect('clients:list')


@login_required
def client_detail(request, pk):
    branch = get_branch(request)
    client = get_object_or_404(Client, pk=pk, branch=branch)
    processes = client.processes.filter(is_active=True)
    return render(request, 'clients/detail.html', {'client': client, 'processes': processes})


@login_required
def export_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    branch = get_branch(request)
    clients = Client.objects.filter(branch=branch, is_active=True)
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    pdf.setFillColor(colors.HexColor('#1a1a2e'))
    pdf.rect(0, height - 60, width, 60, fill=1, stroke=0)
    pdf.setFillColor(colors.white)
    pdf.setFont('Helvetica-Bold', 16)
    pdf.drawString(40, height - 38, 'GB & N.Comin Advocacia - Clientes')
    pdf.setFont('Helvetica', 10)
    y = height - 90
    headers = ['Nome', 'Tipo', 'CPF/CNPJ', 'Telefone', 'Cidade']
    x_positions = [40, 220, 290, 380, 460]
    pdf.setFillColor(colors.HexColor('#e8650a'))
    for i, header in enumerate(headers):
        pdf.drawString(x_positions[i], y, header)
    y -= 20
    pdf.setFillColor(colors.black)
    for client in clients:
        if y < 60:
            pdf.showPage()
            y = height - 60
        pdf.drawString(40, y, client.name[:30])
        pdf.drawString(220, y, client.get_type_display())
        pdf.drawString(290, y, client.cpf_cnpj or '-')
        pdf.drawString(380, y, client.phone)
        pdf.drawString(460, y, client.city_state)
        y -= 18
    pdf.save()
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="clientes.pdf"'
    return response


@login_required
def export_excel(request):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    branch = get_branch(request)
    clients = Client.objects.filter(branch=branch, is_active=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Clientes'
    headers = ['Nome', 'Tipo', 'CPF/CNPJ', 'Telefone', 'Email', 'Cidade', 'Estado', 'Processos']
    header_fill = PatternFill(start_color='E8650A', end_color='E8650A', fill_type='solid')
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for client in clients:
        ws.append([
            client.name, client.get_type_display(), client.cpf_cnpj,
            client.phone, client.email, client.city, client.state, client.process_count,
        ])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="clientes.xlsx"'
    return response


@login_required
def generate_procuracao(request, pk):
    branch = get_branch(request)
    client = get_object_or_404(Client, pk=pk, branch=branch)
    paragraphs = build_procuracao_paragraphs(client, branch)

    pdf_bytes = build_simple_pdf(
        paragraphs,
        client_name=client.name,
        branch_name=branch.name if branch else 'GB & N.Comin Advocacia',
    )
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="procuracao_{client.id}.pdf"'
    return response
