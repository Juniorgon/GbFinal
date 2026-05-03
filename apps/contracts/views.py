from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Sum, Q
from django.utils import timezone
from .models import Contract
from apps.clients.models import Client
from apps.processes.models import Process
import io


def get_branch(request):
    return getattr(request, 'current_branch', request.user.branch)


def get_contract_type_data(request):
    contract_type = request.POST.get('type', 'honorarios')
    custom_type = request.POST.get('custom_type', '').strip()
    if contract_type == 'outro' and not custom_type:
        raise ValueError('Informe o novo tipo de contrato.')
    if contract_type != 'outro':
        custom_type = ''
    return contract_type, custom_type


@login_required
def contract_list(request):
    branch = get_branch(request)
    qs = Contract.objects.filter(branch=branch).select_related('client')
    q = request.GET.get('q', '')
    status = request.GET.get('status', '')
    cliente = request.GET.get('cliente', '')
    judicial = request.GET.get('judicial', '')
    order_by = request.GET.get('order_by', '-created_at')
    order_dir = request.GET.get('order_dir', 'desc')

    if q:
        qs = qs.filter(Q(title__icontains=q) | Q(description__icontains=q) | Q(client__name__icontains=q))
    if status:
        qs = qs.filter(status=status)
    if cliente:
        qs = qs.filter(client_id=cliente)
    if judicial:
        qs = qs.filter(judicial_type=judicial)

    total = qs.count()
    ativos = qs.filter(status='ativo').count()
    concluidos = qs.filter(status='concluido').count()
    suspensos = qs.filter(status='suspenso').count()
    valor_total = qs.aggregate(t=Sum('total_value'))['t'] or 0
    today = timezone.now().date()
    vencendo = qs.filter(status='ativo', end_date__lte=today + __import__('datetime').timedelta(days=30), end_date__gte=today).count()

    clients = Client.objects.filter(branch=branch, is_active=True)

    context = {
        'contracts': qs.order_by(order_by), 'q': q, 'status': status, 'cliente': cliente,
        'judicial': judicial, 'total': total, 'ativos': ativos, 'concluidos': concluidos,
        'suspensos': suspensos, 'valor_total': valor_total, 'vencendo': vencendo, 'clients': clients,
    }
    return render(request, 'contracts/list.html', context)


@login_required
def contract_create(request):
    branch = get_branch(request)
    clients = Client.objects.filter(branch=branch, is_active=True)
    processes = Process.objects.filter(branch=branch, is_active=True)
    if request.method == 'POST':
        try:
            contract_type, custom_type = get_contract_type_data(request)
            c = Contract(
                branch=branch,
                client_id=request.POST.get('client'),
                type=contract_type,
                custom_type=custom_type,
                title=request.POST.get('title', ''),
                description=request.POST.get('description', ''),
                total_value=request.POST.get('total_value') or 0,
                installments=request.POST.get('installments') or 1,
                payment_conditions=request.POST.get('payment_conditions', ''),
                status=request.POST.get('status', 'ativo'),
                judicial_type=request.POST.get('judicial_type', 'judicial'),
                start_date=request.POST.get('start_date'),
                end_date=request.POST.get('end_date'),
                notes=request.POST.get('notes', ''),
                created_by=request.user,
            )
            process_id = request.POST.get('process')
            if process_id:
                c.process_id = process_id
            c.save()
            messages.success(request, 'Contrato criado com sucesso!')
            return redirect('contracts:list')
        except Exception as e:
            messages.error(request, f'Erro: {e}')
    return render(request, 'contracts/form.html', {
        'title': 'Novo Contrato', 'clients': clients, 'processes': processes,
        'type_choices': Contract.TYPE_CHOICES, 'status_choices': Contract.STATUS_CHOICES,
        'judicial_choices': Contract.JUDICIAL_CHOICES,
    })


@login_required
def contract_edit(request, pk):
    branch = get_branch(request)
    contract = get_object_or_404(Contract, pk=pk, branch=branch)
    clients = Client.objects.filter(branch=branch, is_active=True)
    processes = Process.objects.filter(branch=branch, is_active=True)
    if request.method == 'POST':
        try:
            contract_type, custom_type = get_contract_type_data(request)
            contract.client_id = request.POST.get('client')
            contract.type = contract_type
            contract.custom_type = custom_type
            contract.title = request.POST.get('title', '')
            contract.description = request.POST.get('description', '')
            contract.total_value = request.POST.get('total_value') or 0
            contract.installments = request.POST.get('installments') or 1
            contract.payment_conditions = request.POST.get('payment_conditions', '')
            contract.status = request.POST.get('status', 'ativo')
            contract.judicial_type = request.POST.get('judicial_type', 'judicial')
            contract.start_date = request.POST.get('start_date')
            contract.end_date = request.POST.get('end_date')
            contract.notes = request.POST.get('notes', '')
            process_id = request.POST.get('process')
            contract.process_id = process_id if process_id else None
            contract.save()
            messages.success(request, 'Contrato atualizado!')
            return redirect('contracts:list')
        except Exception as e:
            messages.error(request, f'Erro: {e}')
    return render(request, 'contracts/form.html', {
        'title': 'Editar Contrato', 'contract': contract, 'clients': clients, 'processes': processes,
        'type_choices': Contract.TYPE_CHOICES, 'status_choices': Contract.STATUS_CHOICES,
        'judicial_choices': Contract.JUDICIAL_CHOICES,
    })


@login_required
def contract_delete(request, pk):
    branch = get_branch(request)
    contract = get_object_or_404(Contract, pk=pk, branch=branch)
    contract.delete()
    messages.success(request, 'Contrato removido.')
    return redirect('contracts:list')


@login_required
def export_pdf(request):
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    branch = get_branch(request)
    contracts = Contract.objects.filter(branch=branch).select_related('client')
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    w, h = A4
    p.setFillColor(colors.HexColor('#1a1a2e'))
    p.rect(0, h - 60, w, 60, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont('Helvetica-Bold', 14)
    p.drawString(40, h - 38, 'GB & N.Comin Advocacia - Contratos')
    y = h - 90
    p.setFillColor(colors.HexColor('#e8650a'))
    p.setFont('Helvetica-Bold', 9)
    for txt, x in [('Cliente', 40), ('Título', 170), ('Valor', 330), ('Status', 410), ('Vigência', 480)]:
        p.drawString(x, y, txt)
    y -= 20
    p.setFont('Helvetica', 8)
    for c in contracts:
        if y < 60:
            p.showPage(); y = h - 60
        p.setFillColor(colors.black)
        p.drawString(40, y, c.client.name[:20])
        p.drawString(170, y, c.title[:22])
        p.drawString(330, y, f'R$ {c.total_value:,.2f}')
        p.drawString(410, y, c.get_status_display())
        p.drawString(480, y, c.end_date.strftime('%d/%m/%Y') if c.end_date else '-')
        y -= 16
    p.save()
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="contratos.pdf"'
    return response


@login_required
def export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    branch = get_branch(request)
    contracts = Contract.objects.filter(branch=branch).select_related('client')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Contratos'
    headers = ['Cliente', 'Título', 'Tipo', 'Valor Total', 'Parcelas', 'Status', 'Início', 'Término']
    hf = PatternFill(start_color='E8650A', end_color='E8650A', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = hf
        cell.alignment = Alignment(horizontal='center')
    for row, c in enumerate(contracts, 2):
        ws.append([
            c.client.name, c.title, c.type_label, float(c.total_value),
            c.installments, c.get_status_display(),
            c.start_date.strftime('%d/%m/%Y') if c.start_date else '',
            c.end_date.strftime('%d/%m/%Y') if c.end_date else '',
        ])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="contratos.xlsx"'
    return response
